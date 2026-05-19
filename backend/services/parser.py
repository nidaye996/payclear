"""
智能表格解析服务
不依赖固定列名，通过数据格式自动识别字段
支持：Word文档（.docx）、Excel文件（.xlsx/.xls）
处理合并单元格、多行表头等复杂格式
"""
import re
import json
import logging
from typing import Optional, List, Dict, Any, Tuple
from pathlib import Path

logger = logging.getLogger(__name__)

# ==================== 字段识别规则 ====================

def is_id_card(value: str) -> bool:
    """判断是否为身份证号（18位，最后一位可能是X，含校验算法）"""
    if not value:
        return False
    value = str(value).strip().upper()
    if not re.match(r'^\d{17}[\dX]$', value):
        return False
    # 校验位算法
    factors = [7, 9, 10, 5, 8, 4, 2, 1, 6, 3, 7, 9, 10, 5, 8, 4, 2]
    check_chars = '10X98765432'
    try:
        total = sum(int(value[i]) * factors[i] for i in range(17))
        return check_chars[total % 11] == value[17]
    except Exception:
        return False


def is_bank_card(value: str) -> bool:
    """判断是否为银行卡号（16-19位，Luhn算法）"""
    if not value:
        return False
    value = str(value).strip().replace(' ', '').replace('-', '')
    if not re.match(r'^\d{16,19}$', value):
        return False
    # Luhn算法
    digits = [int(d) for d in value]
    odd_digits = digits[-1::-2]
    even_digits = digits[-2::-2]
    total = sum(odd_digits)
    for d in even_digits:
        total += sum(divmod(d * 2, 10))
    return total % 10 == 0


def is_routing_number(value: str) -> bool:
    """判断是否为联行号（12位纯数字）"""
    if not value:
        return False
    value = str(value).strip()
    return bool(re.match(r'^\d{12}$', value))


def is_phone(value: str) -> bool:
    """判断是否为手机号（11位，1开头）"""
    if not value:
        return False
    value = str(value).strip()
    return bool(re.match(r'^1[3-9]\d{9}$', value))


def is_chinese_name(value: str) -> bool:
    """判断是否为中文姓名（2-6个汉字，含少数民族名）"""
    if not value:
        return False
    value = str(value).strip()
    return bool(re.match(r'^[\u4e00-\u9fa5·•]{2,6}$', value))


def is_amount(value: str) -> bool:
    """判断是否为金额"""
    if not value:
        return False
    value = str(value).strip().replace(',', '').replace('，', '')
    return bool(re.match(r'^\d+(\.\d{1,2})?$', value)) and float(value) > 0


def detect_field_type(value: str) -> Optional[str]:
    """
    自动检测单个值的字段类型
    返回：id_card / bank_card / routing_number / phone / name / amount / None
    """
    if not value or str(value).strip() in ('', '-', '/', 'N/A', 'nan', 'None'):
        return None
    value = str(value).strip()

    if is_id_card(value):
        return 'id_card'
    if is_routing_number(value):
        return 'routing_number'
    if is_bank_card(value):
        return 'bank_card'
    if is_phone(value):
        return 'phone'
    if is_chinese_name(value):
        return 'name'
    if is_amount(value):
        return 'amount'
    return None


def infer_column_types(rows: List[List[str]]) -> Dict[int, Dict[str, Any]]:
    """
    通过数据行推断每列的字段类型
    rows: 数据行（已去除表头），每行是一个列表
    返回：{列索引: {'type': 字段类型, 'confidence': 置信度}}
    """
    if not rows:
        return {}

    col_count = max(len(row) for row in rows) if rows else 0
    col_stats: Dict[int, Dict[str, int]] = {}

    for row in rows:
        for col_idx, cell in enumerate(row):
            if col_idx not in col_stats:
                col_stats[col_idx] = {}
            ft = detect_field_type(str(cell) if cell else '')
            if ft:
                col_stats[col_idx][ft] = col_stats[col_idx].get(ft, 0) + 1

    result = {}
    for col_idx, type_counts in col_stats.items():
        if not type_counts:
            continue
        # 取出现次数最多的类型
        best_type = max(type_counts, key=lambda k: type_counts[k])
        total_non_empty = sum(type_counts.values())
        confidence = type_counts[best_type] / total_non_empty if total_non_empty else 0
        if confidence >= 0.5:  # 超过50%的值符合该类型
            result[col_idx] = {
                'type': best_type,
                'confidence': confidence,
                'count': type_counts[best_type]
            }

    return result


def normalize_cell(value: Any) -> str:
    """标准化单元格值"""
    if value is None:
        return ''
    s = str(value).strip()
    # 去除全角空格
    s = s.replace('\u3000', ' ').strip()
    # 清理 nan
    if s.lower() in ('nan', 'none', 'null', '-', '/'):
        return ''
    return s


# ==================== Word 解析 ====================

def parse_word_file(file_path: str) -> List[Dict[str, Any]]:
    """
    解析 Word 文档中的所有表格
    支持合并单元格、多行表头
    返回解析后的工人信息列表
    """
    from docx import Document

    doc = Document(file_path)
    all_workers = []

    for table_idx, table in enumerate(doc.tables):
        try:
            workers = _parse_word_table(table, table_idx)
            all_workers.extend(workers)
        except Exception as e:
            logger.warning(f"解析Word表格[{table_idx}]失败: {e}")

    return all_workers


def _extract_table_rows(table) -> List[List[str]]:
    """
    提取Word表格所有行，直接遍历 XML <w:tc> 元素。
    绕过 python-docx 的 row.cells 在处理垂直合并单元格时会错误丢弃首列的 bug，
    确保每行数据按真实列位置读取。
    水平合并（gridSpan）的单元格重复填充，保持列对齐。
    """
    WNS = 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'
    rows = []
    for row in table.rows:
        cells = []
        tr = row._tr
        for tc in tr.findall(f'{{{WNS}}}tc'):
            # 读取该单元格所有文本
            text = ''.join(t.text or '' for t in tc.iter(f'{{{WNS}}}t'))
            text = text.strip().replace('\n', ' ').replace('\r', '')
            # 处理水平合并：gridSpan > 1 时重复填充
            gs_el = tc.find(f'.//{{{WNS}}}gridSpan')
            span = int(gs_el.get(f'{{{WNS}}}val', '1')) if gs_el is not None else 1
            cells.extend([text] * span)
        rows.append(cells)
    return rows


def _parse_word_table(table, table_idx: int) -> List[Dict[str, Any]]:
    """解析单个Word表格"""
    raw_rows = _extract_table_rows(table)
    if not raw_rows:
        return []

    # 去重（合并单元格会导致相邻行内容相同）
    deduped_rows = []
    prev_row = None
    for row in raw_rows:
        normalized = [normalize_cell(c) for c in row]
        if normalized != prev_row:
            deduped_rows.append(normalized)
            prev_row = normalized

    if len(deduped_rows) < 2:
        return []

    # 找数据起始行：跳过表头（含大量中文、无身份证格式数据的行）
    data_start = _find_data_start_row(deduped_rows)
    if data_start is None:
        return []

    data_rows = deduped_rows[data_start:]
    if not data_rows:
        return []

    # 推断列类型
    col_types = infer_column_types(data_rows)

    # 尝试从表头行补充字段类型
    header_rows = deduped_rows[:data_start]
    col_hints = _extract_header_hints(header_rows, len(data_rows[0]) if data_rows else 0)

    # 表头提示优先，覆盖数据推断
    for col_idx, hint_type in col_hints.items():
        col_types[col_idx] = {'type': hint_type, 'confidence': 0.9}

    # 若表头已明确标出金额列（实发工资），把其他只靠数据推断的金额列删掉
    # 避免出勤天数（31天）被当成工资金额
    hint_amount_cols = {idx for idx, ht in col_hints.items() if ht == 'amount'}
    if hint_amount_cols:
        for idx in list(col_types.keys()):
            if idx not in col_hints and col_types[idx].get('type') == 'amount':
                del col_types[idx]

    return _build_worker_records(data_rows, col_types)


def _find_data_start_row(rows: List[List[str]]) -> Optional[int]:
    """找到数据起始行（第一行含有效身份证号的行）"""
    for i, row in enumerate(rows):
        for cell in row:
            if cell and is_id_card(cell):
                return i
    return None


def _extract_header_hints(header_rows: List[List[str]], col_count: int) -> Dict[int, str]:
    """从表头文字推断列类型"""
    hints = {}
    keywords = {
        '身份证': 'id_card',
        '证件号': 'id_card',
        '银行卡': 'bank_card',
        '卡号': 'bank_card',
        '账号': 'bank_card',
        '收款方账号': 'bank_card',
        '收款账号': 'bank_card',
        '联行号': 'routing_number',
        '行号': 'routing_number',
        '收款银行联行号': 'routing_number',
        '银行联行号': 'routing_number',
        '手机': 'phone',
        '电话': 'phone',
        '联系方式': 'phone',
        '姓名': 'name',
        '收款方名称': 'name',
        '收款人姓名': 'name',
        '收款人': 'name',
        '开户银行': 'bank_name',
        '开户行': 'bank_name',
        '收款方开户行名称': 'bank_name',
        '收款方开户行': 'bank_name',
        '收款方银行名称': 'bank_inst_name',   # 机构全称（支付明细专用列）
        '支付金额': 'amount',
        '金额': 'amount',
        '实发工资': 'net_salary',             # 工资表实发，独立字段
        '实发': 'net_salary',
        '应发工资': 'gross_salary',
        '应发': 'gross_salary',
        '代扣': 'deduction',
        '代缴': 'deduction',
        '出勤工日': 'days_attended',
        '出勤天数': 'days_attended',
        '出勤': 'days_attended',
    }

    # 取最后一行表头（最接近数据的那行）
    for row in reversed(header_rows):
        for col_idx, cell in enumerate(row):
            if col_idx >= col_count:
                break
            for kw, ft in keywords.items():
                if kw in cell and col_idx not in hints:
                    hints[col_idx] = ft
                    break

    return hints


def _build_worker_records(
    data_rows: List[List[str]],
    col_types: Dict[int, Dict[str, Any]],
    require_id_card: bool = True,
) -> List[Dict[str, Any]]:
    """根据列类型映射，构建工人记录列表"""
    workers = []

    for row in data_rows:
        if not any(c.strip() for c in row):
            continue  # 跳过空行

        record: Dict[str, Any] = {}

        for col_idx, type_info in col_types.items():
            if col_idx >= len(row):
                continue
            cell = normalize_cell(row[col_idx])
            if not cell:
                continue

            ft = type_info['type']
            # 每种类型只取第一个找到的值
            if ft == 'id_card' and 'id_card' not in record and is_id_card(cell):
                record['id_card'] = cell.upper()
            elif ft == 'bank_card' and 'bank_card' not in record:
                # 只检查位数，不用 Luhn，避免误杀国内部分银行卡号
                clean = cell.replace(' ', '').replace('-', '')
                if re.match(r'^\d{16,19}$', clean):
                    record['bank_card'] = clean
            elif ft == 'routing_number' and 'routing_number' not in record and is_routing_number(cell):
                record['routing_number'] = cell
            elif ft == 'phone' and 'phone' not in record and is_phone(cell):
                record['phone'] = cell
            elif ft == 'name' and 'name' not in record and is_chinese_name(cell):
                record['name'] = cell
            elif ft == 'bank_name' and 'bank_name' not in record and cell:
                record['bank_name'] = cell
            elif ft == 'bank_inst_name' and 'bank_inst_name' not in record and cell:
                record['bank_inst_name'] = cell
            elif ft == 'amount' and 'amount' not in record and is_amount(cell):
                record['amount'] = float(cell.replace(',', ''))
            elif ft == 'net_salary' and 'net_salary' not in record and is_amount(cell):
                record['net_salary'] = float(cell.replace(',', ''))
            elif ft == 'gross_salary' and 'gross_salary' not in record and is_amount(cell):
                record['gross_salary'] = float(cell.replace(',', ''))
            elif ft == 'deduction' and 'deduction' not in record and is_amount(cell):
                record['deduction'] = float(cell.replace(',', ''))
            elif ft == 'days_attended' and 'days_attended' not in record:
                try:
                    v = float(cell.replace(',', ''))
                    if 0 <= v <= 31:  # 出勤天数不超过31天
                        record['days_attended'] = v
                except ValueError:
                    pass

        if require_id_card:
            # 必须有身份证才算有效记录
            if 'id_card' not in record:
                continue
            # 尝试从相邻列补充未识别的名字
            if 'name' not in record:
                name = _find_name_near_idcard(row, col_types)
                if name:
                    record['name'] = name
            # 尝试从表格文字中提取银行名称（后备方案）
            if 'bank_name' not in record:
                record['bank_name'] = _find_bank_name(row, col_types)
        else:
            # 无需身份证（如考勤表），必须有姓名
            if 'name' not in record:
                continue

        workers.append(record)

    return workers


def _find_name_near_idcard(row: List[str], col_types: Dict[int, Dict[str, Any]]) -> Optional[str]:
    """在行中寻找名字（当智能识别未找到时的后备方案）"""
    for col_idx, cell in enumerate(row):
        cell = normalize_cell(cell)
        if col_idx in col_types:
            continue  # 已识别的列跳过
        if is_chinese_name(cell):
            return cell
    return None


def _find_bank_name(row: List[str], col_types: Dict[int, Dict[str, Any]]) -> Optional[str]:
    """从行中找银行名称（包含"银行"或"信用社"等关键字的列）"""
    bank_keywords = ['银行', '信用社', '农商', '农合', '邮储', '工行', '建行', '农行', '中行', '交行', '村镇银行']
    for col_idx, cell in enumerate(row):
        cell = normalize_cell(cell)
        if not cell:
            continue
        # 已被识别为特定类型的列跳过
        if col_idx in col_types:
            ft = col_types[col_idx]['type']
            if ft in ('id_card', 'bank_card', 'routing_number', 'phone', 'amount'):
                continue
        for kw in bank_keywords:
            if kw in cell and len(cell) > 2:
                return cell
    return None


# ==================== Excel 解析 ====================

def parse_excel_file(file_path: str) -> List[Dict[str, Any]]:
    """
    解析 Excel 文件
    支持 .xlsx 和 .xls 两种格式，自动识别字段
    """
    suffix = Path(file_path).suffix.lower()

    if suffix == '.xls':
        return _parse_xls_file(file_path)

    import openpyxl
    wb = openpyxl.load_workbook(file_path, data_only=True)
    all_workers = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        try:
            workers = _parse_excel_sheet(ws, sheet_name)
            all_workers.extend(workers)
        except Exception as e:
            logger.warning(f"解析Sheet[{sheet_name}]失败: {e}")

    return all_workers


def _parse_xls_file(file_path: str) -> List[Dict[str, Any]]:
    """解析老版 .xls 文件"""
    import xlrd
    wb = xlrd.open_workbook(file_path)
    all_workers = []

    for sheet in wb.sheets():
        try:
            rows = []
            for i in range(sheet.nrows):
                row = []
                for j in range(sheet.ncols):
                    cell = sheet.cell(i, j)
                    # xlrd 类型：0=空,1=文本,2=数字,3=日期,4=布尔,5=错误
                    if cell.ctype == 2:
                        v = cell.value
                        row.append(str(int(v)) if v == int(v) else str(v))
                    elif cell.ctype == 0:
                        row.append('')
                    else:
                        row.append(str(cell.value).strip())
                rows.append(row)

            if not rows:
                continue

            data_start = _find_data_start_row(rows)
            if data_start is None:
                continue

            data_rows = rows[data_start:]
            col_types = infer_column_types(data_rows)
            header_rows = rows[:data_start]
            col_hints = _extract_header_hints(header_rows, len(data_rows[0]) if data_rows else 0)
            for col_idx, hint_type in col_hints.items():
                col_types[col_idx] = {'type': hint_type, 'confidence': 0.9}
            hint_amount_cols = {idx for idx, ht in col_hints.items() if ht == 'amount'}
            if hint_amount_cols:
                for idx in list(col_types.keys()):
                    if idx not in col_hints and col_types[idx].get('type') == 'amount':
                        del col_types[idx]

            all_workers.extend(_build_worker_records(data_rows, col_types))
        except Exception as e:
            logger.warning(f"解析xls Sheet[{sheet.name}]失败: {e}")

    return all_workers


def _parse_excel_sheet(ws, sheet_name: str) -> List[Dict[str, Any]]:
    """解析单个Excel sheet"""
    # 展开合并单元格
    rows = _expand_merged_cells(ws)

    if not rows:
        return []

    # 转字符串
    str_rows = [[normalize_cell(cell) for cell in row] for row in rows]

    # 找数据起始行
    data_start = _find_data_start_row(str_rows)
    if data_start is None:
        return []

    data_rows = str_rows[data_start:]
    col_types = infer_column_types(data_rows)

    # 表头辅助
    header_rows = str_rows[:data_start]
    col_hints = _extract_header_hints(header_rows, len(data_rows[0]) if data_rows else 0)
    for col_idx, hint_type in col_hints.items():
        col_types[col_idx] = {'type': hint_type, 'confidence': 0.9}
    hint_amount_cols = {idx for idx, ht in col_hints.items() if ht == 'amount'}
    if hint_amount_cols:
        for idx in list(col_types.keys()):
            if idx not in col_hints and col_types[idx].get('type') == 'amount':
                del col_types[idx]

    return _build_worker_records(data_rows, col_types)


def _expand_merged_cells(ws) -> List[List[Any]]:
    """展开合并单元格，填充合并区域的值"""
    # 获取合并单元格范围
    merged_ranges = ws.merged_cells.ranges
    merge_map: Dict[Tuple[int, int], Any] = {}

    for merge_range in merged_ranges:
        # 取合并区域左上角的值
        min_row, min_col = merge_range.min_row, merge_range.min_col
        cell_value = ws.cell(row=min_row, column=min_col).value
        for row in range(merge_range.min_row, merge_range.max_row + 1):
            for col in range(merge_range.min_col, merge_range.max_col + 1):
                merge_map[(row, col)] = cell_value

    rows = []
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        cells = []
        for col_idx, cell in enumerate(row, start=1):
            if (row_idx, col_idx) in merge_map:
                cells.append(merge_map[(row_idx, col_idx)])
            else:
                cells.append(cell.value)
        rows.append(cells)

    # 去掉全空的行
    rows = [r for r in rows if any(c is not None and str(c).strip() not in ('', 'None') for c in r)]
    return rows


# ==================== 考勤表专用解析 ====================

def parse_attendance_file(file_path: str) -> List[Dict[str, Any]]:
    """
    解析考勤表（Excel格式）：不含身份证，以姓名+出勤工日为主要字段。
    策略：找到"出勤工日"或"出勤天数"等列，提取姓名和出勤天数。
    """
    suffix = Path(file_path).suffix.lower()
    if suffix not in ('.xlsx', '.xls'):
        return []

    try:
        if suffix == '.xls':
            import xlrd
            wb = xlrd.open_workbook(file_path)
            sheets_data = []
            for sheet in wb.sheets():
                rows = []
                for i in range(sheet.nrows):
                    row = []
                    for j in range(sheet.ncols):
                        cell = sheet.cell(i, j)
                        if cell.ctype == 2:
                            v = cell.value
                            row.append(str(int(v)) if v == int(v) else str(v))
                        elif cell.ctype == 0:
                            row.append('')
                        else:
                            row.append(str(cell.value).strip())
                    rows.append(row)
                sheets_data.append(rows)
        else:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheets_data = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                raw = _expand_merged_cells(ws)
                rows = [[normalize_cell(c) for c in row] for row in raw]
                sheets_data.append(rows)
    except Exception as e:
        logger.warning(f"考勤表打开失败 {file_path}: {e}")
        return []

    workers = []
    for rows in sheets_data:
        if not rows:
            continue

        # 找包含"出勤"关键字的表头行
        header_row_idx = None
        name_col = None
        days_col = None
        for i, row in enumerate(rows):
            for j, cell in enumerate(row):
                c = normalize_cell(str(cell)) if cell else ''
                if '出勤' in c and ('工日' in c or '天数' in c or '天' in c):
                    days_col = j
                    header_row_idx = i
                elif '姓名' in c and name_col is None:
                    name_col = j
            if header_row_idx is not None and name_col is not None and days_col is not None:
                break

        if header_row_idx is None or name_col is None or days_col is None:
            # 后备：用智能识别
            header_rows_hint = rows[:5] if len(rows) > 5 else rows
            col_hints = _extract_header_hints(header_rows_hint, max(len(r) for r in rows) if rows else 0)
            # 找 days_attended 列
            for col_idx, ft in col_hints.items():
                if ft == 'days_attended':
                    days_col = col_idx
                elif ft == 'name':
                    name_col = col_idx
            # 找数据起始行（有中文名字的第一行）
            data_start = 0
            for i, row in enumerate(rows):
                if name_col is not None and name_col < len(row):
                    cell = normalize_cell(str(row[name_col]))
                    if is_chinese_name(cell):
                        data_start = i
                        break
            if name_col is None or days_col is None:
                continue
            data_rows = rows[data_start:]
        else:
            data_rows = rows[header_row_idx + 1:]

        for row in data_rows:
            if not row:
                continue
            name_cell = normalize_cell(str(row[name_col])) if name_col < len(row) else ''
            days_cell = normalize_cell(str(row[days_col])) if days_col < len(row) else ''

            if not is_chinese_name(name_cell):
                continue
            try:
                days = float(days_cell.replace(',', ''))
                if not (0 <= days <= 31):
                    continue
            except (ValueError, TypeError):
                continue

            workers.append({'name': name_cell, 'days_attended': days})

    # 去重（以姓名为key，保留第一条）
    seen_names = set()
    unique = []
    for w in workers:
        if w['name'] not in seen_names:
            seen_names.add(w['name'])
            unique.append(w)
    return unique


# ==================== 站班会数据专用解析 ====================

def parse_station_meeting_file(file_path: str) -> Dict[str, Any]:
    """
    解析站班会列表数据（Excel格式）。
    返回：
      {
        'attendance_by_name': {姓名: 出勤天数},   # 按日期去重后的出勤次数
        'duplicate_days': {姓名: [重复日期列表]},  # 同一人同一天出现多次的情况
      }
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        logger.warning(f"站班会文件打开失败 {file_path}: {e}")
        return {'attendance_by_name': {}, 'duplicate_days': {}}

    if not rows:
        return {'attendance_by_name': {}, 'duplicate_days': {}}

    # 找表头行，定位"施工日期"、"班组负责人"、"其他施工人员"列
    header_idx = None
    date_col = leader_col = others_col = None
    for i, row in enumerate(rows):
        for j, cell in enumerate(row):
            v = str(cell or '').strip()
            if '施工日期' in v:
                date_col = j; header_idx = i
            elif '班组负责人' in v:
                leader_col = j
            elif '其他施工人员' in v:
                others_col = j

    if header_idx is None or date_col is None:
        return {'attendance_by_name': {}, 'duplicate_days': {}}

    # name -> {date -> count}
    name_date_count: Dict[str, Dict[str, int]] = {}

    def add_name(name: str, date_str: str):
        name = name.strip().strip('（）()').strip()
        # 去掉括号内备注，如"张益铭（退场）"→"张益铭"
        import re
        name = re.sub(r'[（(][^）)]*[）)]', '', name).strip()
        if not name or len(name) < 2:
            return
        if name not in name_date_count:
            name_date_count[name] = {}
        name_date_count[name][date_str] = name_date_count[name].get(date_str, 0) + 1

    for row in rows[header_idx + 1:]:
        if not row:
            continue
        date_val = row[date_col] if date_col < len(row) else None
        if date_val is None:
            continue
        # 统一日期为字符串
        if hasattr(date_val, 'strftime'):
            date_str = date_val.strftime('%Y-%m-%d')
        else:
            date_str = str(date_val).strip()[:10]
        if not date_str or date_str == 'None':
            continue

        # 班组负责人
        if leader_col is not None and leader_col < len(row):
            for name in str(row[leader_col] or '').split(','):
                add_name(name, date_str)

        # 其他施工人员
        if others_col is not None and others_col < len(row):
            for name in str(row[others_col] or '').split(','):
                add_name(name, date_str)

    attendance_by_name: Dict[str, int] = {}
    duplicate_days: Dict[str, list] = {}

    for name, date_count in name_date_count.items():
        dups = [d for d, cnt in date_count.items() if cnt > 1]
        if dups:
            duplicate_days[name] = dups
        attendance_by_name[name] = len(date_count)  # 按日期去重后的天数

    return {
        'attendance_by_name': attendance_by_name,
        'duplicate_days': duplicate_days,
    }


# ==================== 统一入口 ====================

def parse_file(file_path: str, file_type_hint: str = '') -> Dict[str, Any]:
    """
    统一解析入口
    file_path: 文件路径
    file_type_hint: 文件类型提示（可选）
    返回：{'workers': [...], 'error': None/str, 'raw_count': int}
    """
    path = Path(file_path)
    suffix = path.suffix.lower()

    try:
        # 考勤表：无身份证，用专用解析器
        if file_type_hint == '考勤表':
            workers = parse_attendance_file(file_path)
            return {
                'workers': workers,
                'error': None,
                'raw_count': len(workers)
            }

        # 站班会数据：专用解析器，返回格式特殊
        if file_type_hint == '站班会数据':
            result = parse_station_meeting_file(file_path)
            return {
                'workers': [],  # 站班会不产生 worker 列表
                'station_meeting': result,
                'error': None,
                'raw_count': len(result.get('attendance_by_name', {}))
            }

        if suffix == '.docx':
            workers = parse_word_file(file_path)
        elif suffix in ('.xlsx', '.xls'):
            workers = parse_excel_file(file_path)
        else:
            return {
                'workers': [],
                'error': f'不支持的文件格式: {suffix}',
                'raw_count': 0
            }

        # 去重（同一份表里身份证重复的情况）
        seen_ids = set()
        unique_workers = []
        for w in workers:
            idc = w.get('id_card', '')
            if idc and idc not in seen_ids:
                seen_ids.add(idc)
                unique_workers.append(w)

        return {
            'workers': unique_workers,
            'error': None,
            'raw_count': len(workers)
        }

    except Exception as e:
        logger.error(f"解析文件失败 {file_path}: {e}", exc_info=True)
        return {
            'workers': [],
            'error': str(e),
            'raw_count': 0
        }
